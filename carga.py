from openpyxl import load_workbook
import pandas_datareader.data as web
import datetime

wb = load_workbook('datos_variables.xlsm')
ws = wb.active
fechas = ws['A238':'A2266']
datos_a_cargar = []
for x in fechas:
    for fecha in x:
        datos_a_cargar.append([fecha.value.strftime("%Y-%m-%d")])
vxx = ws['G238':'G2266']
i = 0
for x in vxx:
    for valor in x:
        datos_a_cargar[i].append(valor.value)
    i += 1





start = datetime.datetime(2010, 1, 1)
end = datetime.datetime.now()

df_vix = web.DataReader("^VIX", 'yahoo', start, end)

dates_vix =[]
for x in range(len(df_vix)):
    newdate = str(df_vix.index[x])
    newdate = newdate[0:10]
    dates_vix.append(newdate)

df_vix['dates'] = dates_vix

df_vxx = web.DataReader("VXX", 'yahoo', start, end)

dates_vxx =[]
for x in range(len(df_vxx)):
    newdate = str(df_vxx.index[x])
    newdate = newdate[0:10]
    dates_vxx.append(newdate)

df_vxx['dates'] = dates_vxx
for date,vxx in zip(dates_vxx,df_vxx['Close']):
    datos_a_cargar.append([date, vxx])
vxx = [(date,value) for date, value in datos_a_cargar]
vix = []
for date,v_vix in zip(dates_vix,df_vix['Close']):
    vix.append((date, v_vix))
