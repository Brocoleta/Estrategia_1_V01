from carga import vxx, vix
import numpy as np
import pandas as pd
import matplotlib.pyplot as pp
import matplotlib.dates
import pandas_datareader.data as web
from parametros import *
import datetime
from openpyxl import load_workbook,Workbook
from openpyxl.chart import LineChart,Reference
from openpyxl.chart.axis import DateAxis




from os import path




class Equity:
    def __init__(self,capital_disponible,ino,primer_rebote,avu,stop_prof,stop_loss,dias_stop_loss,vxx,vix,exp_stop_loss,pos_exposicion):
        super().__init__()
        self.capital_disponible = capital_disponible
        self.capital = capital_disponible
        self.ino = ino
        self.primer_rebote = primer_rebote
        self.pos_exposicion = pos_exposicion

        self.avu = avu

        self.primer_rebote_bool = False

        self.stop_prof = stop_prof
        self.stop_loss = stop_loss
        self.dias_stop_loss = dias_stop_loss
        self.vxx = vxx
        self.vix = vix
        self.exp_stop_loss = exp_stop_loss
        self.df = self.curva_equity(self.vxx)
        self.max_drawdawn()
        self.excel()




    def curva_equity(self,fecha_inicio='2010-01-01'):

        vxx = np.array(self.vxx)
        vix = np.array(self.vix)

        self.stop_loss_bool = False
        # valor de entrada
        self.valor_inicial = float(vxx[0][1])
        # valor anterior al actual de ahi se va actualizando
        self.valor_anterior = float(vxx[0][1])
        # valor donde la curva empieza a subir de nuevo
        self.valor_rebote = float(vxx[0][1])
        self.valor_alza = float(vxx[0][1])
        # posiciones inicales
        self.posicion = round((self.capital_disponible*self.ino)/self.valor_inicial)
        # me dice si estoy en la bolsa o no, cuando el vix es > stop_proof se actualiza
        self.estoy_en_la_bolsa = False
        self.posicion_relativa = 0
        self.valor_primer_rebote = 0
        self.valor_segundo_rebote = 0
        self.valor_tercer_rebote = 0
        # utilidades
        self.utilidades = 0
        self.utilidades_realizadas = 0
        self.posicion_primer_rebote = 0
        self.posicion_segundo_rebote = 0
        self.posicion_tercer_rebote = 0
        self.drawdowns = []
        self.drawdown = 0.
        self.max_seen = float(vxx[0][1])
        self.posiciones_total = []



        self.posiciones = [(self.posicion, self.valor_inicial)]
        self.expociciones = []
        self.stop_loss_dia = 0


        def equity(close):
            close[0][1] = float(close[0][1])
            close[1][1] = float(close[1][1])



            # VALOR VXX
            self.valor_actual = float(close[0][1])
            # close[1][1] es el valor del VIX, esto revisa que no este por debajo de 13

            if self.stop_loss_bool:
                self.stop_loss_dia += 1



            if self.stop_loss_dia == self.dias_stop_loss:
                self.stop_loss_bool = False



            if not self.stop_loss_bool:


                if float(close[1][1]) <= float(self.stop_prof) and not self.stop_loss_bool:

                    self.equity = 0
                    for posicion, rebote in self.posiciones:
                        self.equity += posicion*(rebote - float(self.valor_actual))
                    self.capital_disponible += self.equity
                    self.equity = self.capital_disponible
                    self.posiciones = []
                    self.posicion = 0


                    self.estoy_en_la_bolsa = False

                elif float(close[1][1]) > float(self.stop_prof) and not self.stop_loss_bool:
                    # esto pasa solo cuando vuelve a entrar a la bolsa, actualiza la posicion y el valor principal
                    if not self.estoy_en_la_bolsa:

                        self.estoy_en_la_bolsa = True
                        self.posicion = round((self.capital_disponible*self.ino)/self.valor_actual)
                        self.posiciones = [(self.posicion, self.valor_actual)]

                        self.valor_inicial = self.valor_actual
                        self.valor_alza = self.valor_actual
                        self.valor_rebote = self.valor_actual




                    # esto pasa cuando la curva cae debajo del valor donde empezo a subir(valor rebote)
                    if self.valor_actual < self.valor_alza:

                        if self.valor_anterior < self.valor_actual:
                            # es cuando empieza a subir denuevo, actualiza el valor de rebote(donde empieza a subir) y reinica los parametros que dicen si ya ha subido 10 25 50

                            self.valor_alza = self.valor_anterior
                            self.valor_rebote = self.valor_anterior


                    self.equity = 0
                    for posicion, rebote in self.posiciones:
                        self.equity += posicion*(rebote - float(self.valor_actual))
                    self.equity += self.capital_disponible



                    # Estos if revisan si ha subido un 10,25,50, si ese es el caso aumenta la posicion y el valor de la utilidad con ese valor de vxx
                    #if (self.valor_actual - self.valor_rebote) > (self.valor_rebote * self.tercer_rebote) and not self.tercer_rebote_bool:
                        # Estos True son para decir que ya subio ese porcentaje una vez, para que no lo haga tod el rato
                        #self.tercer_rebote_bool = True
                        #self.nueva_pos = round((self.equity*self.avu)/self.valor_actual)
                        #self.posiciones.append((self.nueva_pos, self.valor_actual))

                        # aumenta la posicion en equity*avu/vxx

                        # aumenta las utilidades en (aumento de posiciones)*(valor de entrada - valor actual) <- esto me esta pasando que me da negativo, debe ser cuando entra a la bolsa y empieza a subir, no se si esta bien el calculo


                    #if (self.valor_actual - self.valor_rebote) > (self.valor_rebote * self.segundo_rebote) and not self.segundo_rebote_bool:
                        #self.segundo_rebote_bool = True
                        #self.nueva_pos = round((self.equity*self.avu)/self.valor_actual)
                        #self.posiciones.append((self.nueva_pos, self.valor_actual))
                    self.posicion_x = 0
                    for p,v in self.posiciones:
                        self.posicion_x += p
                    expocicion = self.valor_actual*self.posicion_x/self.equity





                    if (self.valor_actual - self.valor_rebote) > (self.valor_rebote* self.primer_rebote) and not self.primer_rebote_bool and expocicion<self.pos_exposicion:
                        self.valor_rebote = self.valor_actual
                        self.nueva_pos = round((self.equity*self.avu)/self.valor_actual)
                        self.posiciones.append((self.nueva_pos, self.valor_actual))
                        self.posicion_x = 0
                        for p,v in self.posiciones:
                            self.posicion_x += p
                        expocicion = self.valor_actual*self.posicion_x/self.equity
                        if expocicion>self.pos_exposicion:
                            self.posicion_x -= self.nueva_pos
                            self.posiciones.pop()
                            self.nueva_pos = round((self.pos_exposicion*self.equity/self.valor_actual)-self.posicion_x)
                            self.posiciones.append((self.nueva_pos, self.valor_actual))



                self.max_seen = max(self.max_seen, self.equity)
                self.drawdowns.append(1 - self.equity / self.max_seen)
                #self.datos.append((close[0][0],self.equity,self.capital_disponible,1 - self.equity / self.max_seen))

            else:
                self.equity = self.capital_disponible
                self.drawdowns.append(0)
                #self.datos.append((close[0][0],self.equity,self.capital_disponible,0))
            self.posicion_x = 0
            for p,v in self.posiciones:
                self.posicion_x += p
            expocicion = self.valor_actual*self.posicion_x/self.equity
            self.expociciones.append(expocicion)
            if (len(self.drawdowns) > 0 and self.drawdowns[-1] > self.stop_loss and not self.stop_loss_bool) or (expocicion > self.exp_stop_loss and not self.stop_loss_bool):

                self.stop_loss_bool = True
                self.equity = 0
                for posicion, rebote in self.posiciones:
                    self.equity += posicion*(rebote - float(self.valor_actual))

                self.capital_disponible += self.equity
                self.equity = self.capital_disponible
                self.posiciones = []
                self.posicion = 0
                self.stop_loss_dia = 1
                self.max_seen = 0

                self.estoy_en_la_bolsa = False






            self.valor_anterior = self.valor_actual
            self.posicion_x = 0
            for p,v in self.posiciones:
                self.posicion_x += p
            self.posiciones_total.append(self.posicion_x)

            #datos.append((f'equity = {equity}',f'fecha = {close[0][0]}',f'vxx = {close[0][1]}',f'vix = {close[1][1]}'))

            #actualiza el valor valor_anterior

            #retorna el valor del equity en ese punto
            return self.equity
        # esto aplica la funcion del para calcular el equity en cada punto
        self.equitys = np.array([equity(close) for close in zip(vxx,vix)])
        # todo lo siguiente no tiene relevancia, es para graficar
        self.fechas = np.array([close[0][0] for close in zip(vxx,vix)])

        df = pd.DataFrame()

        df['datetime'] = self.fechas
        df['equitys'] = self.equitys
        df.datetime = pd.to_datetime(df.datetime)
        df.set_index('datetime', inplace=True)
        self.drawdowns = np.array([d*100 for d in self.drawdowns])
        #pp.plot(self.drawdowns)
        #pp.show()

        return self.equitys,self.drawdowns, self.expociciones, self.posiciones_total
    def df_func(self,vxx):
        self.fechas = np.array([close[0] for close in vxx])

        df = pd.DataFrame()

        df['datetime'] = self.fechas
        df['equitys'] = self.equitys
        df.datetime = pd.to_datetime(df.datetime)
        df.set_index('datetime', inplace=True)





        #pp.plot(self.drawdowns)
        #pp.show()

        return df
    def calc_MDD(self, equitys):
      df = pd.Series(equitys, name="nw").to_frame()

      max_peaks_idx = df.nw.expanding(min_periods=1).apply(lambda x: x.idxmax(),raw=False).fillna(0).astype(int)
      df['max_peaks_idx'] = pd.Series(max_peaks_idx).to_frame()

      nw_peaks = pd.Series(df.nw.iloc[max_peaks_idx.values].values, index=df.nw.index)

      df['dd'] = ((df.nw-nw_peaks)/nw_peaks)
      df['mdd'] = df.groupby('max_peaks_idx').dd.apply(lambda x: x.expanding(min_periods=1).apply(lambda y: y.min(),raw=False)).fillna(0)


      return list(df['mdd'])

    def max_drawdawn(self):

        equity = []
        equitys = []
        t = False
        i = 0
        for e,p in zip(self.equitys, self.posiciones_total):
            if p != 0:

                equity.append(e)
            else:
                if len(equity) > 0:
                    equity.append(e)

                    equitys.extend(self.calc_MDD(equity))
                    equity = []
                else:
                    equitys.append(0)
            i += 1
        if len(equity) > 0:
            equity.append(e)

            equitys.extend(self.calc_MDD(equity))


        drawdawns = np.flip(np.array(equitys))
        self.max_drawdawns = []

        maximo = drawdawns[0]

        for d in drawdawns:
            if d != 0:
                if d < maximo:
                    maximo = d

            else:
                maximo = 0
            self.max_drawdawns.append(maximo)
        self.max_drawdawns = np.flip(np.array(self.max_drawdawns))







    def excel(self):
        df = self.df
        vxx = self.vxx
        vix = self.vix
        if path.exists("yahooxslx.xlsx"):
            wb = load_workbook("yahooxslx.xlsx")

            # grab the active worksheet
            ws = wb.create_sheet()
        else:
            wb = Workbook()

            # grab the active worksheet
            ws = wb.active



        ws['A1'] = 'Fecha'
        ws['B1'] = 'VXX'
        ws['C1'] = 'VIX'
        ws['D1'] = 'Equity'
        ws['E1'] = '%Max Drawdown'
        ws['F1'] = '%Expocicion'
        ws['G1'] = 'Posicion'
        ws['H1'] = 'Max Drawdown'
        ws['I1'] = 'Max Equity'
        ws['J1'] = 'Parametros'
        ws['K1'] = 'Capital Inicial'
        ws['L1'] = '%Rebote'
        ws['M1'] = 'Aumento Volumen Umbral'
        ws['N1'] = '%Inicial'
        ws['O1'] = 'Stop Profit VIX'
        ws['P1'] = 'Stop Loss'
        ws['Q1'] = 'Stop Loss Expocicion'
        ws['R1'] = 'Stop Loss Dias'
        ws['S1'] = 'Stop Aumento posicion-exposicion'
        ws['K2'] = self.capital
        ws['L2'] = self.primer_rebote*100
        ws['M2'] = self.avu*100
        ws['N2'] = self.ino*100
        ws['O2'] = self.stop_prof
        ws['P2'] = self.stop_loss*100
        ws['Q2'] = self.exp_stop_loss*100
        ws['R2'] = self.dias_stop_loss
        ws['I2'] = np.max(df[0])
        ws['H2'] = -np.min(self.max_drawdawns)
        ws['S2'] = self.pos_exposicion



        i = 2

        for vxx,vix,equity,dw,exp,pos in zip(vxx,vix,df[0],self.max_drawdawns,df[2],df[3]):
            ws[f'A{i}'] = vxx[0]
            ws[f'B{i}'] = float(vxx[1])
            ws[f'C{i}'] = float(vix[1])
            ws[f'D{i}'] = float(equity)
            ws[f'E{i}'] = "{0:0.1f}".format(abs(dw*100))
            ws[f'F{i}'] = "{0:0.1f}".format(exp*100)
            ws[f'G{i}'] = int(pos)

            i += 1

        chart = LineChart()
        chart.title = "Curva Equity"
        chart.style = 13
        chart.x_axis.title = 'Fecha'
        chart.y_axis.title = 'Equity'
        chart.height = 10 # default is 7.5
        chart.width = 20
        chart.y_axis.crossAx = 500
        chart.x_axis = DateAxis(crossAx=100)
        chart.x_axis.number_format ='yyyy/mm/dd'
        chart.x_axis.majorTimeUnit = "days"
        data = Reference(ws, min_col=4, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        dates = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.set_categories(dates)
        ws.add_chart(chart, "J8")

        wb.save('yahooxslx.xlsx')
